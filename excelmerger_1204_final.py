# -*- coding: utf-8 -*-
"""
Created on Thu Nov 12 17:01:16 2020

@author: rudolf.kenyeres
"""
import openpyxl, os
from openpyxl.utils import range_boundaries
import pyjokes

excelFiles = []
for filename in os.listdir('.'):
    if filename.endswith('.xlsx'):
        excelFiles.append(filename)
excelFiles.sort(key=str.lower)

#Intitializes workbooks
print('Searching data files...')
min_col, min_row, max_col, max_row = range_boundaries("A:CA")

  
keyWord= [item for item in input("Enter the examined keywords, separate by commas : ").split(",")]
KeywordIndex  = [int(item) for item in input("Enter the examined column index, which should be containing the keyword(s): ").split(",")]
minimum = int(input("Enter the minimum value: "))
MinimumIndex= [int(item) for item in input("Enter the examined column index, which should be greather than the minimum: ").split(",")]


template = openpyxl.load_workbook('output.xlsx')
templatews = template.active

for i in excelFiles:  
    wb = openpyxl.load_workbook(i, data_only=True) 
    ws = wb.active
    for word in keyWord:                
        #Does the actual thing
        for wordindex in KeywordIndex:
            for minindex in MinimumIndex:
                for row in ws.iter_rows():
                    if row[wordindex].value == str(word) and row[minindex].value != None and int(row[minindex].value) >= minimum:
                        templatews.append(cell.value for cell in row[min_col-1:max_col])
template.save('output.xlsx')
#df = pd.DataFrame(templatews.values)

    #Saves the results
print('All done! Have fun! Here is a joke:')
print(pyjokes.get_joke())
print("Press Enter to continue ...")
input()